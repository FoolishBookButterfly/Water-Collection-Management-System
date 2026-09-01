import openpyxl
import json
from datetime import datetime

# Load the workbook
wb = openpyxl.load_workbook('record.xlsx')

print(f"Sheet names: {wb.sheetnames}")

# Get first sheet (occupants)
ws = wb.active
print(f"\nActive sheet: {ws.title}")

# Extract headers
headers = []
for col_idx in range(1, ws.max_column + 1):
    cell = ws.cell(2, col_idx)
    headers.append(str(cell.value).strip() if cell.value else f"Col{col_idx}")

# Extract collection dates from headers (starting from column 7)
collection_dates = []
for col_idx in range(7, ws.max_column + 1):
    header = headers[col_idx - 1]
    if header and header != "Remarks":
        # Parse date
        try:
            # header might be like "2026-08-11 00:00:00"
            date_obj = datetime.fromisoformat(header.split(' ')[0])
            collection_dates.append({
                'date': date_obj.strftime('%Y-%m-%d'),
                'displayDate': date_obj.strftime('%B %d, %Y'),
                'columnIndex': col_idx
            })
        except:
            collection_dates.append({
                'date': header,
                'displayDate': header,
                'columnIndex': col_idx
            })

print(f"Collection dates: {len(collection_dates)}")
for d in collection_dates:
    print(f"  {d}")

# Extract occupants
occupants = []
for row_idx in range(3, 145):  # Rows 3 to 144 (before TOTAL)
    surname = ws.cell(row_idx, 2).value
    if not surname or not str(surname).strip():
        continue
    
    first_name = ws.cell(row_idx, 3).value or ""
    course = ws.cell(row_idx, 4).value or ""
    room = ws.cell(row_idx, 5).value or ""
    remarks = ws.cell(row_idx, 6).value or ""
    
    # Parse payment status for each date
    payments = {}
    for date_info in collection_dates:
        col_idx = date_info['columnIndex']
        payment_cell = ws.cell(row_idx, col_idx).value or ""
        payment_str = str(payment_cell).strip()
        
        # Determine status
        if payment_str == "/":
            status = "paid"
        elif payment_str == "excused":
            status = "excused"
        elif "kulang" in payment_str.lower():
            status = "partial"
        elif payment_str == "":
            status = "unpaid"
        else:
            status = "unpaid"
        
        payments[date_info['date']] = status
    
    occupant = {
        'id': f"occ_{row_idx}",
        'no': row_idx - 2,
        'surname': str(surname).strip(),
        'firstName': str(first_name).strip(),
        'course': str(course).strip(),
        'room': str(room).strip(),
        'remarks': str(remarks).strip(),
        'payments': payments
    }
    occupants.append(occupant)

print(f"\n\nTotal occupants: {len(occupants)}")
print(f"First occupant: {occupants[0]}")

# Extract totals from TOTAL row
totals_row = 145
total_data = {}
for date_info in collection_dates:
    col_idx = date_info['columnIndex']
    total_cell = ws.cell(totals_row, col_idx).value
    total_data[date_info['date']] = int(total_cell) if total_cell else 0

print(f"\nTotals: {total_data}")

# For now, assume no purchases sheet - will check
purchases = []
if len(wb.sheetnames) > 1:
    print(f"\nChecking sheet 2: {wb.sheetnames[1]}")
    ws2 = wb[wb.sheetnames[1]]
    # Look for purchase data
    for row_idx in range(1, min(20, ws2.max_row + 1)):
        col1 = ws2.cell(row_idx, 1).value
        col2 = ws2.cell(row_idx, 2).value
        col3 = ws2.cell(row_idx, 3).value
        print(f"  Row {row_idx}: {col1} | {col2} | {col3}")

# Create output JSON
output = {
    'occupants': occupants,
    'collectionDates': [d['date'] for d in collection_dates],
    'totals': total_data,
    'purchases': purchases,
    'currentBalance': sum(total_data.values()) if total_data else 0
}

# Save to file
with open('seed_data.json', 'w') as f:
    json.dump(output, f, indent=2)

print("\n\nData saved to seed_data.json")
print(f"Total records: {len(occupants)}")
print(f"Collection dates: {len(collection_dates)}")
print(f"Current total collected: {sum(total_data.values())}")
