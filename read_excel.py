import openpyxl
import json

# Load the workbook
wb = openpyxl.load_workbook('record.xlsx')
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"Dimensions: {ws.max_row} rows x {ws.max_column} cols")
print("\n=== ALL DATA ===\n")

# Print all cells
for row_idx in range(1, ws.max_row + 1):
    row_data = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row_idx, col_idx)
        value = cell.value
        row_data.append(str(value) if value is not None else "")
    print(f"Row {row_idx}: {row_data}")
