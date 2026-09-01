import openpyxl
import json
from datetime import datetime

# Load the workbook
wb = openpyxl.load_workbook('record.xlsx')
ws = wb.active

# Find the header row (should contain "No.", "Surname", "First Name", etc.)
header_row = None
for row_idx in range(1, 50):
    cell_value = ws.cell(row_idx, 2).value  # Check 2nd column
    if cell_value and str(cell_value).strip().upper() in ['SURNAME', 'NO.', 'NO']:
        header_row = row_idx
        break

print(f"Header row: {header_row}")

# Extract header
if header_row:
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(header_row, col_idx)
        headers.append(str(cell.value).strip() if cell.value else "")
    print(f"Headers: {headers}")

# Find the data rows and total row
data_rows = []
total_row_idx = None

for row_idx in range(header_row + 1 if header_row else 1, ws.max_row + 1):
    # Check if this is the TOTAL row
    remarks_cell = ws.cell(row_idx, 6).value
    if remarks_cell and str(remarks_cell).strip().upper() == "TOTAL":
        total_row_idx = row_idx
        print(f"Found TOTAL row at: {row_idx}")
        break
    
    # Check if row has any data
    surname = ws.cell(row_idx, 2).value
    if surname and str(surname).strip():
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            row_data.append(str(cell.value).strip() if cell.value else "")
        data_rows.append((row_idx, row_data))

print(f"\nTotal occupant rows found: {len(data_rows)}")
print(f"\nFirst 5 occupants:")
for i, (idx, row) in enumerate(data_rows[:5]):
    print(f"  {idx}: {row}")

# If TOTAL row found, extract totals
if total_row_idx:
    print(f"\nTOTAL row data (row {total_row_idx}):")
    total_data = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(total_row_idx, col_idx)
        total_data.append(str(cell.value) if cell.value else "")
    print(f"  {total_data}")

# Look for purchases section
print(f"\n\nLooking for purchases section after row {total_row_idx if total_row_idx else ws.max_row}...")
if total_row_idx:
    for row_idx in range(total_row_idx + 1, min(total_row_idx + 20, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, 5):
            cell = ws.cell(row_idx, col_idx)
            row_data.append(str(cell.value) if cell.value else "")
        if any(row_data):
            print(f"  Row {row_idx}: {row_data}")
